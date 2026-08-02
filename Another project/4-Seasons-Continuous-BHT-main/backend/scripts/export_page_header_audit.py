#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PAGE_DEFINITIONS = [
    ("screener-demographics", "Screener & Demographics", [r"^S\d+", r"^Q\d+", r"\bSEC\b", r"\bAge\b", r"\bmarital\b"]),
    ("category-questions", "Category Questions", [r"_QC\d*", r"\bcategory question",]),
    ("awareness-usage", "Awareness & Usage", [r"_BAU\d*", r"\bawareness\b", r"\bseen or heard\b"]),
    ("purchase-behavior", "Purchase Behavior", [r"_PB\d*", r"\bpurchase behavior\b", r"\bbuy\b"]),
    ("brand-imagery", "Brand Imagery", [r"_QBI\d*", r"\bbrand imagery\b", r"\bbrand that\b"]),
    ("flavour-flex-section", "Flavour/Flex Section", [r"_FQ\d*", r"_QFS\d*", r"_QFSB\d*", r"_QFW\d*", r"\bflavou?r\b", r"\bflex\b"]),
    ("campaign-check", "Campaign Check", [r"_A\d*", r"\bcampaign check\b", r"\badvert", r"\bcampaign\b"]),
    ("video-ad-section", "Video Ad Section", [r"\bvideo ad section\b", r"\bvideo\b"]),
    ("other-metrics", "Other Metrics", []),
]
PAGE_REGEXES = [(page_id, title, [re.compile(pattern, re.I) for pattern in patterns]) for page_id, title, patterns in PAGE_DEFINITIONS]
PAGE_TITLE_BY_ID = {page_id: title for page_id, title, _patterns in PAGE_DEFINITIONS}
PAGE_SORT_ORDER = {
    "Overview": 10,
    "Awareness - Brand Awareness": 20,
    "Awareness - Ad Awareness": 21,
    "Awareness - Media Source": 22,
    "Awareness - Usage": 23,
    "Screener & Demographics": 30,
    "Category Questions": 40,
    "Awareness & Usage - Raw Questions": 50,
    "Purchase Behavior": 60,
    "Brand Imagery": 70,
    "Flavour": 80,
    "Flex": 81,
    "Flavour/Flex Section": 82,
    "Campaign Check": 90,
    "Video Ad Section": 100,
    "Other Metrics": 110,
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
OVERVIEW_DIMENSIONS = [
    ("Overview", "", "Region", "overview_dimension", ["Region"], [], "respondent_dims", "Dashboard overview card label 'Region'"),
    ("Overview", "", "Income", "overview_dimension", ["D3"], [], "respondent_dims", "Dashboard overview card label 'Income' is backed by respondent_dims.D3"),
    ("Overview", "", "Gender", "overview_dimension", ["Gender"], [], "respondent_dims", "Dashboard overview card label 'Gender'"),
    ("Overview", "", "Age", "overview_dimension", ["Age"], [], "respondent_dims", "Dashboard overview card label 'Age'"),
    ("Overview", "", "SEC", "overview_dimension", ["SEC"], [], "respondent_dims", "Dashboard overview card label 'SEC'"),
    ("Overview", "", "Week", "overview_dimension", ["Week"], [], "respondent_dims", "Dashboard overview card label 'Week'"),
]


@dataclass
class QuestionRow:
    category: str
    question: str
    question_label: str


def compact(value: object) -> str:
    return " ".join(str(value or "").split())


def classify_question(question: str, question_label: str) -> str:
    text = f"{question or ''} {question_label or ''}"
    for page_id, _title, regexes in PAGE_REGEXES:
        if regexes and any(regex.search(text) for regex in regexes):
            return page_id
    return "other-metrics"


def extract_question_family_code(question_code: str) -> str:
    code = compact(question_code)
    parts = code.split("_")
    for part in parts:
        if re.fullmatch(r"(?:BAU\d+[A-Z]?|PB\d+[A-Z]?|QC\d+[A-Z]?|A\d+)", part, re.I):
            return part.upper()
    return ""


def group_key_for_question_code(question_code: str, page_id: str) -> str:
    compact_code = compact(question_code)
    family_code = extract_question_family_code(compact_code)
    if family_code and page_id != "brand-imagery":
        return family_code
    return re.sub(r"_\d+$", "", compact_code)


def strip_trailing_option(label: str) -> str:
    return re.sub(r"\s*[:;,\-]?\s*\([^)]*\)\s*$", "", compact(label)).strip()


def unique_preserve(values: Iterable[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = compact(value)
        if not text or text in seen:
            continue
        seen.add(text)
        out.append(text)
    return out


def detect_flavour_flex_subpage(question: str, question_label: str) -> str:
    text = f"{question} {question_label}"
    if re.search(r"_FQ\d*|\bflavou?r\b", text, re.I):
        return "Flavour"
    if re.search(r"_QFS\d*|_QFSB\d*|_QFW\d*|\bflex\b", text, re.I):
        return "Flex"
    return "Flavour/Flex Section"


def fetch_categories(con: duckdb.DuckDBPyConnection) -> list[str]:
    rows = con.execute(
        """
        SELECT DISTINCT CAST(category AS VARCHAR) AS category
        FROM (
          SELECT category FROM respondent_dims
          UNION ALL
          SELECT category FROM responses_fact
        )
        WHERE category IS NOT NULL
        ORDER BY category
        """
    ).fetchall()
    return [compact(row[0]) for row in rows if compact(row[0])]


def fetch_questions(con: duckdb.DuckDBPyConnection, category: str) -> list[QuestionRow]:
    tables = {
        row[0]
        for row in con.execute(
            """
            SELECT table_name
            FROM information_schema.tables
            WHERE table_schema = 'main'
            """
        ).fetchall()
    }
    if "question_catalog" in tables:
        sql = """
            SELECT
              CAST(category AS VARCHAR) AS category,
              CAST(question AS VARCHAR) AS question,
              COALESCE(CAST(question_label AS VARCHAR), CAST(question AS VARCHAR)) AS question_label
            FROM question_catalog
            WHERE CAST(category AS VARCHAR) = ?
            ORDER BY question
        """
    else:
        sql = """
            SELECT DISTINCT
              CAST(category AS VARCHAR) AS category,
              CAST(question AS VARCHAR) AS question,
              COALESCE(CAST(question_label AS VARCHAR), CAST(question AS VARCHAR)) AS question_label
            FROM responses_fact
            WHERE CAST(category AS VARCHAR) = ?
            ORDER BY question
        """
    return [QuestionRow(*row) for row in con.execute(sql, [category]).fetchall()]


def fetch_matching_questions(con: duckdb.DuckDBPyConnection, category: str, regexes: list[str]) -> tuple[list[str], list[str]]:
    questions = fetch_questions(con, category)
    compiled = [re.compile(pattern, re.I) for pattern in regexes]
    matched = [row for row in questions if any(regex.search(row.question) for regex in compiled)]
    return (
        unique_preserve(row.question for row in matched),
        unique_preserve(row.question_label for row in matched),
    )


def build_grouped_question_rows(category: str, questions: list[QuestionRow]) -> list[dict[str, object]]:
    grouped: dict[tuple[str, str], list[QuestionRow]] = defaultdict(list)
    order_index: dict[tuple[str, str], int] = {}

    for index, row in enumerate(questions):
        page_id = classify_question(row.question, row.question_label)
        key = group_key_for_question_code(row.question, page_id)
        composite_key = (page_id, key)
        if composite_key not in order_index:
            order_index[composite_key] = index
        grouped[composite_key].append(row)

    out: list[dict[str, object]] = []
    for (page_id, _group_key), items in sorted(grouped.items(), key=lambda item: order_index[item[0]]):
        first = items[0]
        is_grouped = len(items) > 1
        header_label = strip_trailing_option(first.question_label) if is_grouped else compact(first.question_label or first.question)
        if not header_label:
            header_label = compact(first.question)
        page_label = PAGE_TITLE_BY_ID.get(page_id, page_id)
        subpage = ""
        if page_id == "awareness-usage":
            page_label = "Awareness & Usage - Raw Questions"
        elif page_id == "flavour-flex-section":
            subpage = detect_flavour_flex_subpage(first.question, first.question_label)
            page_label = subpage

        out.append(
            {
                "page": page_label,
                "subpage": subpage,
                "header_label": header_label,
                "header_type": "grouped_question",
                "source_variable_names": unique_preserve(item.question for item in items),
                "source_question_labels": unique_preserve(item.question_label for item in items),
                "source_table": "question_catalog" if items else "responses_fact",
                "note": f"Grouped from {len(items)} question code(s)" if is_grouped else "Single question header",
            }
        )
    return out


def build_fixed_metric_rows(con: duckdb.DuckDBPyConnection, category: str) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    metric_configs = [
        ("Awareness - Brand Awareness", "Brand TOM", [r"(?i)(^|_)BAU1A$"], "responses_fact", "Derived awareness column"),
        ("Awareness - Brand Awareness", "Brand SPONT", [r"(?i)(^|_)BAU1B(?:_[0-9]+)?$"], "responses_fact", "Derived awareness column"),
        ("Awareness - Brand Awareness", "AIDED", [r"(?i)(^|_)BAU2(?:_[0-9]+)?$"], "responses_fact", "Derived awareness column"),
        ("Awareness - Brand Awareness", "Total Awareness", [r"(?i)(^|_)BAU1A$", r"(?i)(^|_)BAU1B(?:_[0-9]+)?$", r"(?i)(^|_)BAU2(?:_[0-9]+)?$"], "derived", "Derived as Brand TOM + Brand SPONT + AIDED"),
        ("Awareness - Ad Awareness", "AD TOM", [r"(?i)(^|_)BAU1C$"], "responses_fact", "Derived ad awareness column"),
        ("Awareness - Ad Awareness", "AD SPONT", [r"(?i)(^|_)BAU1D(?:_[0-9]+)?$"], "responses_fact", "Derived ad awareness column"),
        ("Awareness - Ad Awareness", "AIDED AD", [r"(?i)(^|_)BAU3(?:_[0-9]+)?$"], "responses_fact", "Derived ad awareness column"),
        ("Awareness - Ad Awareness", "Total Ad Awareness", [r"(?i)(^|_)BAU1C$", r"(?i)(^|_)BAU1D(?:_[0-9]+)?$", r"(?i)(^|_)BAU3(?:_[0-9]+)?$"], "derived", "Derived as AD TOM + AD SPONT + AIDED AD"),
        ("Awareness - Media Source", "Media Source", [r"(?i)(^|_)BAU4(?:_[0-9]+)?$"], "responses_fact", "Media source page"),
        ("Awareness - Usage", "Ever Consumed", [r"(?i)(^|_)BAU5A(?:_[0-9]+)?$"], "responses_fact", "Usage page metric"),
        ("Awareness - Usage", "Last 3 Months", [r"(?i)(^|_)BAU5C(?:_[0-9]+)?$"], "responses_fact", "Usage page metric"),
        ("Awareness - Usage", "Last 1 Month", [r"(?i)(^|_)BAU6A(?:_[0-9]+)?$"], "responses_fact", "Usage page metric"),
        ("Awareness - Usage", "Last 7 Days", [r"(?i)(^|_)BAU6B(?:_[0-9]+)?$"], "responses_fact", "Usage page metric"),
        ("Awareness - Usage", "Most Often Used", [r"(?i)(^|_)BAU6C$"], "responses_fact", "Usage page metric"),
        ("Awareness - Usage", "Prefrence", [r"(?i)(^|_)BAU8$"], "responses_fact", "Usage page metric"),
    ]
    for page, header_label, regexes, source_table, note in metric_configs:
        variable_names, question_labels = fetch_matching_questions(con, category, regexes)
        rows.append(
            {
                "page": page,
                "subpage": "",
                "header_label": header_label,
                "header_type": "derived_metric",
                "source_variable_names": variable_names,
                "source_question_labels": question_labels,
                "source_table": source_table,
                "note": note,
            }
        )
    return rows


def safe_sheet_title(raw_title: str, used: set[str]) -> str:
    base = re.sub(r"[:\\/?*\[\]]", "_", compact(raw_title)) or "Sheet"
    base = base[:31]
    candidate = base
    counter = 2
    while candidate in used:
        suffix = f"_{counter}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        "A": 18,
        "B": 28,
        "C": 20,
        "D": 38,
        "E": 20,
        "F": 52,
        "G": 12,
        "H": 60,
        "I": 18,
        "J": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_summary_sheet(wb: Workbook, db_path: Path, categories: list[str], status_note: str | None = None) -> None:
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ["Field", "Value"],
        ["Database path", str(db_path)],
        ["Generated at", datetime.now().isoformat(timespec="seconds")],
        ["Database exists", "Yes" if db_path.exists() else "No"],
        ["Category count", len(categories)],
        ["Categories", ", ".join(categories)],
    ]
    if status_note:
        rows.append(["Status", status_note])
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 120
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_rows_for_category(con: duckdb.DuckDBPyConnection, category: str) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for page, subpage, header_label, header_type, source_variables, source_labels, source_table, note in OVERVIEW_DIMENSIONS:
        rows.append(
            {
                "page": page,
                "subpage": subpage,
                "header_label": header_label,
                "header_type": header_type,
                "source_variable_names": source_variables,
                "source_question_labels": source_labels,
                "source_table": source_table,
                "note": note,
            }
        )

    rows.extend(build_fixed_metric_rows(con, category))
    rows.extend(build_grouped_question_rows(category, fetch_questions(con, category)))
    rows.sort(key=lambda item: (PAGE_SORT_ORDER.get(str(item["page"]), 999), str(item["page"]), str(item["header_label"])))
    return rows


def export_workbook(db_path: Path, output_path: Path) -> None:
    wb = Workbook()
    used_sheet_titles = {"Summary"}

    if not db_path.exists():
        write_summary_sheet(wb, db_path, [], status_note="DuckDB file was not found in the workspace. No category sheets could be generated.")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return

    con = duckdb.connect(str(db_path), read_only=True)
    try:
        categories = fetch_categories(con)
        write_summary_sheet(wb, db_path, categories)
        for category in categories:
            ws = wb.create_sheet(title=safe_sheet_title(category, used_sheet_titles))
            ws.append(
                [
                    "Category",
                    "Page",
                    "Subpage/View",
                    "Header Label",
                    "Header Type",
                    "Source Variable Name(s)",
                    "Variable Count",
                    "Source Question Label(s)",
                    "Source Table",
                    "Notes",
                ]
            )
            for row in build_rows_for_category(con, category):
                variable_names = unique_preserve(row["source_variable_names"])
                question_labels = unique_preserve(row["source_question_labels"])
                ws.append(
                    [
                        category,
                        row["page"],
                        row["subpage"],
                        row["header_label"],
                        row["header_type"],
                        "\n".join(variable_names),
                        len(variable_names),
                        "\n".join(question_labels),
                        row["source_table"],
                        row["note"],
                    ]
                )
            style_sheet(ws)
    finally:
        con.close()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Export page/header-to-variable audit workbook from DuckDB.")
    parser.add_argument("--db", required=True, help="Path to the DuckDB file")
    parser.add_argument("--out", required=True, help="Output xlsx path")
    args = parser.parse_args()

    export_workbook(Path(args.db).resolve(), Path(args.out).resolve())
    print(f"Wrote workbook to {Path(args.out).resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
