#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


SKIP_SHEETS = {"Summary", "Sheet1"}


def compact(value: object) -> str:
    return " ".join(str(value or "").split())


def split_lines(value: object) -> list[str]:
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n")
    out: list[str] = []
    seen: set[str] = set()
    for part in text.split("\n"):
        item = compact(part)
        if not item:
            continue
        if item.upper().endswith("_OTH"):
            continue
        if item in seen:
            continue
        seen.add(item)
        out.append(item)
    return out


def workbook_to_payload(workbook_path: Path) -> dict[str, object]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    categories: dict[str, dict[str, object]] = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        rows: list[dict[str, object]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            category = compact(row[0] if len(row) > 0 else "")
            page = compact(row[1] if len(row) > 1 else "")
            subpage = compact(row[2] if len(row) > 2 else "")
            header_label = compact(row[3] if len(row) > 3 else "")
            header_type = compact(row[4] if len(row) > 4 else "")
            variable_names = split_lines(row[5] if len(row) > 5 else "")
            question_labels = split_lines(row[7] if len(row) > 7 else "")
            source_table = compact(row[8] if len(row) > 8 else "")
            notes = compact(row[9] if len(row) > 9 else "")

            if not category:
                category = compact(sheet_name)
            if not page or not header_label:
                continue

            rows.append(
                {
                    "page": page,
                    "subpage": subpage,
                    "headerLabel": header_label,
                    "headerType": header_type,
                    "sourceVariableNames": variable_names,
                    "sourceQuestionLabels": question_labels,
                    "sourceTable": source_table,
                    "notes": notes,
                }
            )

        categories[compact(sheet_name)] = {
            "category": compact(sheet_name),
            "rows": rows,
        }

    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceWorkbook": str(workbook_path.resolve()),
        "categories": categories,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export header-audit workbook category sheets to JSON metadata.")
    parser.add_argument("--workbook", required=True, type=Path, help="Path to the source audit workbook")
    parser.add_argument("--output", required=True, type=Path, help="Path to write the JSON metadata")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    payload = workbook_to_payload(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote metadata to {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
