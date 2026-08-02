#!/usr/bin/env python3
from __future__ import annotations

import argparse
import html
import re
import warnings
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
import pyreadstat


CATEGORY_PREFIX_TO_NAME = {
    "BC": "Breakfast_Cereal",
    "BL": "Bleach",
    "CM": "Condiment_Mixes",
    "DH": "Dry_Hair_Care",
    "EO": "Edible_Oil",
    "ML": "Malt_Beverage",
    "N": "Noodles",
    "SK": "Snacks_Products",
    "TC": "Toilet_Cleaner",
    "TP": "Toothpaste",
    "WH": "Wet_Hair_Care",
}
PAGE_DEFINITIONS = [
    ("screener-demographics", "Screener & Demographics", [r"^S\d+", r"^Q\d+", r"\bSEC\b", r"\bAge\b", r"\bmarital\b"]),
    ("category-questions", "Category Questions", [r"_QC\d*", r"\bcategory question"]),
    ("awareness-usage", "Awareness & Usage", [r"_BAU\d*", r"\bawareness\b", r"\bseen or heard\b"]),
    ("purchase-behavior", "Purchase Behavior", [r"_PB\d*", r"\bpurchase behavior\b", r"\bbuy\b"]),
    ("brand-imagery", "Brand Imagery", [r"_QBI\d*", r"\bbrand imagery\b", r"\bbrand that\b"]),
    ("flavour-flex-section", "Flavour/Flex Section", [r"_FQ\d*", r"_QFS\d*", r"_QFSB\d*", r"_QFW\d*", r"\bflavou?r\b", r"\bflex\b"]),
    ("campaign-check", "Campaign Check", [r"_A\d*", r"\bcampaign check\b", r"\badvert", r"\bcampaign\b"]),
    ("video-ad-section", "Video Ad Section", [r"\bvideo ad section\b", r"\bvideo\b"]),
    ("other-metrics", "Other Metrics", []),
]
PAGE_REGEXES = [(page_id, title, [re.compile(pattern, re.I) for pattern in patterns]) for page_id, title, patterns in PAGE_DEFINITIONS]
PAGE_TITLE_BY_ID = {page_id: title for page_id, title, _patterns in PAGE_DEFINITIONS}
PAGE_SORT_ORDER = {
    "Overview": 10,
    "Awareness - Brand Awareness": 20,
    "Awareness - Ad Awareness": 21,
    "Awareness - Media Source": 22,
    "Awareness - Usage": 23,
    "Screener & Demographics": 30,
    "Category Questions": 40,
    "Awareness & Usage - Raw Questions": 50,
    "Purchase Behavior": 60,
    "Brand Imagery": 70,
    "Flavour": 80,
    "Flex": 81,
    "Flavour/Flex Section": 82,
    "Campaign Check": 90,
    "Video Ad Section": 100,
    "Other Metrics": 110,
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
KEEP_SHEETS = {"Summary", "Sheet1"}
HTML_RE = re.compile(r"<[^>]+>")
SPACE_RE = re.compile(r"\s+")


@dataclass
class VariableRow:
    name: str
    label: str


def compact(value: object) -> str:
    return " ".join(str(value or "").split())


def clean_label(value: object) -> str:
    text = html.unescape(str(value or ""))
    text = HTML_RE.sub(" ", text)
    text = SPACE_RE.sub(" ", text).strip()
    return text


def classify_question(question: str, question_label: str) -> str:
    text = f"{question or ''} {question_label or ''}"
    for page_id, _title, regexes in PAGE_REGEXES:
        if regexes and any(regex.search(text) for regex in regexes):
            return page_id
    return "other-metrics"


def extract_question_family_code(question_code: str) -> str:
    parts = compact(question_code).split("_")
    for part in parts:
        if re.fullmatch(r"(?:BAU\d+[A-Z]?|PB\d+[A-Z]?|QC\d+(?:\.\d+)?|A\d+|QBI(?:\.\d+)?)", part, re.I):
            return part.upper()
    return ""


def group_key_for_question_code(question_code: str, page_id: str) -> str:
    compact_code = compact(question_code)
    family_code = extract_question_family_code(compact_code)
    if family_code and page_id != "brand-imagery":
      return family_code
    return re.sub(r"_\d+$", "", compact_code)


def strip_trailing_option(label: str) -> str:
    return re.sub(r"\s*[:;,\-]?\s*\([^)]*\)\s*$", "", compact(label)).strip()


def unique_preserve(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = compact(value)
        if not text or text in seen:
            continue
        seen.add(text)
        out.append(text)
    return out


def load_sheet_rows(ws) -> list[VariableRow]:
    rows: list[VariableRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = compact(row[0])
        label = clean_label(row[1])
        if not name:
            continue
        rows.append(VariableRow(name=name, label=label))
    return rows


def load_spss_label_map(spss_path: Path) -> dict[str, str]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        _df, meta = pyreadstat.read_sav(str(spss_path), metadataonly=True)

    label_map: dict[str, str] = {}
    for name, label in zip(meta.column_names, meta.column_labels):
        cleaned_name = compact(name)
        cleaned_label = clean_label(label)
        if cleaned_name:
            label_map[cleaned_name] = cleaned_label
    return label_map


def apply_spss_labels(rows: list[VariableRow], label_map: dict[str, str]) -> list[VariableRow]:
    out: list[VariableRow] = []
    for row in rows:
        out.append(
            VariableRow(
                name=row.name,
                label=label_map.get(row.name, row.label),
            )
        )
    return out


def is_other_variable(name: str) -> bool:
    return compact(name).upper().endswith("_OTH")


def infer_prefix(name: str) -> str | None:
    match = re.match(r"^(BC|BL|CM|DH|EO|ML|N|SK|TC|TP|WH)(?:[_.]|$)", name, re.I)
    return match.group(1).upper() if match else None


def get_category_rows(rows: list[VariableRow], prefix: str) -> list[VariableRow]:
    return [row for row in rows if infer_prefix(row.name) == prefix and not is_other_variable(row.name)]


def pick_best_label(items: list[VariableRow], family_code: str) -> str:
    if not items:
        return ""
    informative = []
    family_pattern = re.compile(re.escape(family_code), re.I) if family_code else None
    for item in items:
        label = compact(item.label)
        if not label:
            continue
        score = 0
        if family_pattern and family_pattern.search(label):
            score += 50
        if len(label) >= 20:
            score += 10
        if ":" in label:
            score += 5
        informative.append((score, len(label), label))
    if informative:
        informative.sort(key=lambda entry: (entry[0], entry[1], entry[2]))
        return informative[-1][2]
    return items[0].label or items[0].name


def detect_flavour_flex_subpage(question: str, question_label: str) -> str:
    text = f"{question} {question_label}"
    if re.search(r"_FQ\d*|\bflavou?r\b", text, re.I):
        return "Flavour"
    if re.search(r"_QFS\d*|_QFSB\d*|_QFW\d*|\bflex\b", text, re.I):
        return "Flex"
    return "Flavour/Flex Section"


def build_grouped_question_rows(rows: list[VariableRow]) -> list[dict[str, object]]:
    grouped: dict[tuple[str, str], list[VariableRow]] = defaultdict(list)
    order_index: dict[tuple[str, str], int] = {}
    for index, row in enumerate(rows):
        page_id = classify_question(row.name, row.label)
        group_key = group_key_for_question_code(row.name, page_id)
        composite = (page_id, group_key)
        if composite not in order_index:
            order_index[composite] = index
        grouped[composite].append(row)

    out: list[dict[str, object]] = []
    for (page_id, group_key), items in sorted(grouped.items(), key=lambda item: order_index[item[0]]):
        representative = pick_best_label(items, group_key)
        page = PAGE_TITLE_BY_ID.get(page_id, page_id)
        subpage = ""
        if page_id == "awareness-usage":
            page = "Awareness & Usage - Raw Questions"
        elif page_id == "flavour-flex-section":
            subpage = detect_flavour_flex_subpage(items[0].name, representative)
            page = subpage

        out.append(
            {
                "page": page,
                "subpage": subpage,
                "header_label": strip_trailing_option(representative) or representative or items[0].name,
                "header_type": "grouped_question",
                "source_variable_names": unique_preserve([item.name for item in items]),
                "source_question_labels": unique_preserve([item.label for item in items]),
                "source_table": "SPSS metadata",
                "note": f"Grouped from {len(items)} variable(s)",
            }
        )
    return out


def build_overview_rows(rows: list[VariableRow]) -> list[dict[str, object]]:
    available = {row.name.lower(): row for row in rows}
    source_candidates = {
        "Region": ["Region", "City_1"],
        "Income": ["D3", "d3_q"],
        "Gender": ["Gender", "S4_1"],
        "Age": ["Age", "Age_cal"],
        "SEC": ["SEC", "N_QC1.1"],
        "Week": ["Week"],
    }
    out = []
    for header, candidates in source_candidates.items():
        matches = [available[name.lower()].name for name in candidates if name.lower() in available]
        labels = [available[name.lower()].label for name in candidates if name.lower() in available and available[name.lower()].label]
        out.append(
            {
                "page": "Overview",
                "subpage": "",
                "header_label": header,
                "header_type": "overview_dimension",
                "source_variable_names": matches or candidates[:1],
                "source_question_labels": unique_preserve(labels),
                "source_table": "SPSS metadata",
                "note": f"Dashboard overview card label '{header}'",
            }
        )
    return out


def build_fixed_metric_rows(rows: list[VariableRow], prefix: str) -> list[dict[str, object]]:
    metric_configs = [
        ("Awareness - Brand Awareness", "Brand TOM", [fr"^{prefix}_BAU1a$"], "Derived awareness column"),
        ("Awareness - Brand Awareness", "Brand SPONT", [fr"^{prefix}_BAU1b(?:_\d+)?$"], "Derived awareness column"),
        ("Awareness - Brand Awareness", "AIDED", [fr"^{prefix}_BAU2(?:_\d+)?$"], "Derived awareness column"),
        ("Awareness - Brand Awareness", "Total Awareness", [fr"^{prefix}_BAU1a$", fr"^{prefix}_BAU1b(?:_\d+)?$", fr"^{prefix}_BAU2(?:_\d+)?$"], "Derived as Brand TOM + Brand SPONT + AIDED"),
        ("Awareness - Ad Awareness", "AD TOM", [fr"^{prefix}_BAU1c$"], "Derived ad awareness column"),
        ("Awareness - Ad Awareness", "AD SPONT", [fr"^{prefix}_BAU1d(?:_\d+)?$"], "Derived ad awareness column"),
        ("Awareness - Ad Awareness", "AIDED AD", [fr"^{prefix}_BAU3(?:_\d+)?$"], "Derived ad awareness column"),
        ("Awareness - Ad Awareness", "Total Ad Awareness", [fr"^{prefix}_BAU1c$", fr"^{prefix}_BAU1d(?:_\d+)?$", fr"^{prefix}_BAU3(?:_\d+)?$"], "Derived as AD TOM + AD SPONT + AIDED AD"),
        ("Awareness - Media Source", "Media Source", [fr"^{prefix}_BAU4(?:\.\d+)?(?:_\d+)?$"], "Media source page"),
        ("Awareness - Usage", "Ever Consumed", [fr"^{prefix}_BAU5a(?:_\d+)?$"], "Usage page metric"),
        ("Awareness - Usage", "Last 3 Months", [fr"^{prefix}_BAU5c(?:_\d+)?$"], "Usage page metric"),
        ("Awareness - Usage", "Last 1 Month", [fr"^{prefix}_BAU6a(?:_\d+)?$"], "Usage page metric"),
        ("Awareness - Usage", "Last 7 Days", [fr"^{prefix}_BAU6b(?:_\d+)?$"], "Usage page metric"),
        ("Awareness - Usage", "Most Often Used", [fr"^{prefix}_BAU6c$"], "Usage page metric"),
        ("Awareness - Usage", "Prefrence", [fr"^{prefix}_BAU8(?:\.\d+)?$"], "Usage page metric"),
    ]
    out = []
    for page, header, patterns, note in metric_configs:
        regexes = [re.compile(pattern, re.I) for pattern in patterns]
        matched = [row for row in rows if not is_other_variable(row.name) and any(regex.search(row.name) for regex in regexes)]
        out.append(
            {
                "page": page,
                "subpage": "",
                "header_label": header,
                "header_type": "derived_metric",
                "source_variable_names": unique_preserve([row.name for row in matched]),
                "source_question_labels": unique_preserve([row.label for row in matched]),
                "source_table": "SPSS metadata",
                "note": note,
            }
        )
    return out


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        "A": 18,
        "B": 28,
        "C": 20,
        "D": 38,
        "E": 20,
        "F": 52,
        "G": 12,
        "H": 60,
        "I": 18,
        "J": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def update_summary(ws, categories: list[str], spss_path: Path | None = None) -> None:
    rows = list(ws.iter_rows(min_row=1, max_col=2))
    index_by_field = {}
    for row_index, row in enumerate(rows, start=1):
        field = compact(row[0].value)
        if field:
            index_by_field[field] = row_index
    ws.cell(index_by_field["Category count"], 2).value = len(categories)
    ws.cell(index_by_field["Categories"], 2).value = ", ".join(categories)
    if spss_path is not None:
        if "SPSS path" in index_by_field:
            ws.cell(index_by_field["SPSS path"], 2).value = str(spss_path)
        else:
            ws.append(["SPSS path", str(spss_path)])
    if "Status" in index_by_field:
        ws.cell(index_by_field["Status"], 2).value = "Category audit rebuilt from Sheet1 variable inventory using SPSS labels; *_OTH excluded."
    else:
        ws.append(["Status", "Category audit rebuilt from Sheet1 variable inventory using SPSS labels; *_OTH excluded."])


def build_workbook(workbook_path: Path, spss_path: Path, output_path: Path | None = None) -> Path:
    wb = load_workbook(workbook_path)
    if "Sheet1" not in wb.sheetnames:
        raise ValueError("Sheet1 not found in workbook.")

    source_rows = load_sheet_rows(wb["Sheet1"])
    spss_label_map = load_spss_label_map(spss_path)
    source_rows = apply_spss_labels(source_rows, spss_label_map)
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in KEEP_SHEETS:
            del wb[sheet_name]

    categories = [CATEGORY_PREFIX_TO_NAME[prefix] for prefix in CATEGORY_PREFIX_TO_NAME]
    update_summary(wb["Summary"], categories, spss_path=spss_path)

    for prefix, category_name in CATEGORY_PREFIX_TO_NAME.items():
        category_rows = get_category_rows(source_rows, prefix)
        if not category_rows:
            continue
        ws = wb.create_sheet(title=category_name)
        ws.append(
            [
                "Category",
                "Page",
                "Subpage/View",
                "Header Label",
                "Header Type",
                "Source Variable Name(s)",
                "Variable Count",
                "Source Question Label(s)",
                "Source Table",
                "Notes",
            ]
        )
        audit_rows = build_overview_rows(source_rows)
        audit_rows.extend(build_fixed_metric_rows(category_rows, prefix))
        audit_rows.extend(build_grouped_question_rows(category_rows))
        audit_rows.sort(key=lambda item: (PAGE_SORT_ORDER.get(str(item["page"]), 999), str(item["page"]), str(item["header_label"])))

        for row in audit_rows:
            variable_names = unique_preserve(list(row["source_variable_names"]))
            source_labels = unique_preserve(list(row["source_question_labels"]))
            ws.append(
                [
                    category_name,
                    row["page"],
                    row["subpage"],
                    row["header_label"],
                    row["header_type"],
                    "\n".join(variable_names),
                    len(variable_names),
                    "\n".join(source_labels),
                    row["source_table"],
                    row["note"],
                ]
            )
        style_sheet(ws)

    target_path = output_path or workbook_path
    wb.save(target_path)
    return target_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Build current audit workbook from Sheet1 variable inventory and SPSS labels.")
    parser.add_argument("--workbook", required=True, help="Path to current_duckdb_header_variable_audit.xlsx")
    parser.add_argument("--spss", required=True, help="Path to the SPSS .sav file")
    parser.add_argument("--out", help="Optional output workbook path")
    args = parser.parse_args()
    output_path = Path(args.out).resolve() if args.out else None
    written_path = build_workbook(Path(args.workbook).resolve(), Path(args.spss).resolve(), output_path=output_path)
    print(f"Updated workbook {written_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
