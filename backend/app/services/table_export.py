"""Excel writer for analysis tables.

The supplied full-data workbook is a reporting specification, not a source of
manually-maintained results.  This module copies its sheets, formatting and
layout, then replaces bases and table cells with the calculated values.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any


def _normalise(value: Any) -> str:
    return " ".join(str(value or "").replace("\u2019", "'").lower().split())


def _table_for_heading(tables: list[dict[str, Any]], heading: Any) -> dict[str, Any] | None:
    heading_key = _normalise(heading)
    for table in tables:
        if _normalise(table["title"]) == heading_key:
            return table
    # Titles in the layout can be cleaned-up versions of the BackEnd wording.
    for table in tables:
        if _normalise(table["title"]).split(".", 1)[0] == heading_key.split(".", 1)[0]:
            return table
    return None


def _columns(table: dict[str, Any]) -> dict[str, dict[str, Any]]:
    """Return every displayed demographic column keyed by its visible label."""
    result: dict[str, dict[str, Any]] = {"total": {"base": table["base"], "counts": {r["label"]: r["count"] for r in table["rows"]}}}
    for cut in table.get("cuts", []):
        for group in cut.get("groups", []):
            # The template's headers are display labels, so that is the stable
            # contract between template and calculation engine.
            result.setdefault(_normalise(group["label"]), {"base": group["base"], "counts": group["counts"]})
    return result


def _write_template_sheet(sheet, tables: list[dict[str, Any]], percentage: bool) -> None:
    """Fill an existing report sheet without changing its layout or styling."""
    for row in range(1, sheet.max_row + 1):
        # Each block's base row is the reliable anchor in the supplied template.
        if _normalise(sheet.cell(row, 2).value) != "total" or _normalise(sheet.cell(row, 3).value) not in {"base total", ""}:
            continue
        table = _table_for_heading(tables, sheet.cell(row, 1).value)
        if not table:
            continue
        columns = _columns(table)
        for column in range(4, sheet.max_column + 1):
            header = "total" if column == 4 else _normalise(sheet.cell(4, column).value)
            series = columns.get(header)
            if series:
                sheet.cell(row, column).value = series["base"]
        # Response lines run until the next block title/blank area.  Preserve
        # the report's established answer order and show zeroes where needed.
        for answer_row in range(row + 1, sheet.max_row + 1):
            if answer_row > row + 1 and sheet.cell(answer_row, 1).value is not None:
                break
            response = sheet.cell(answer_row, 2).value
            if response is None:
                continue
            response_key = _normalise(response)
            for column in range(4, sheet.max_column + 1):
                header = "total" if column == 4 else _normalise(sheet.cell(4, column).value)
                series = columns.get(header)
                if not series:
                    continue
                count = next((value for label, value in series["counts"].items() if _normalise(label) == response_key), 0)
                sheet.cell(answer_row, column).value = round(count / series["base"] * 100, 1) if percentage and series["base"] else count


def write_analysis_workbook(path: Path, category: str, tables: list[dict[str, Any]], template_path: Path | None = None) -> None:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill

    if template_path and template_path.exists():
        workbook = load_workbook(template_path)
        percentage_sheet = workbook[category.title()] if category.title() in workbook.sheetnames else None
        count_name = f"{category.title()}_count"
        count_sheet = workbook[count_name] if count_name in workbook.sheetnames else None
        if percentage_sheet:
            _write_template_sheet(percentage_sheet, tables, percentage=True)
        if count_sheet:
            _write_template_sheet(count_sheet, tables, percentage=False)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        return

    # Fallback remains useful for a category with no supplied report template.
    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("Home")
    summary.append([f"BHT Tracker - {category.title()} analysis"])
    summary.append(["Table", "Question", "Base"])
    for table in tables:
        summary.append([table["title"], table["question"], table["base"]])
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1565C0")
    for index, table in enumerate(tables, start=1):
        sheet = workbook.create_sheet(f"Table {index}")
        sheet.append([table["title"]])
        sheet.append([table["question"]])
        sheet.append(["Response", "Total N", "Total %"])
        sheet.append(["Base", table["base"], 100 if table["base"] else 0])
        for row in table["rows"]:
            sheet.append([row["label"], row["count"], row["pct"]])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
