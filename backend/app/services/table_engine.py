"""Reusable survey-table calculations used by both the API and Excel exports."""

from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path
from typing import Any, Iterable, Sequence


NOODLES_TABLES = [
    ("N_BAU1A", "Top-of-mind brand awareness"),
    ("N_BAU1B", "Other spontaneous brand awareness"),
    ("N_BAU1C", "Top-of-mind advertising awareness"),
    ("N_BAU1D", "Other advertising awareness"),
    ("N_BAU2", "Prompted brand awareness"),
    ("N_BAU5A", "Ever consumed"),
    ("N_BAU5C", "Consumed in the last 3 months"),
    ("N_BAU6A", "Consumed in the last month"),
    ("N_BAU6B", "Consumed in the last 7 days"),
    ("N_BAU6C", "Main brand"),
    ("N_BAU8", "Preferred brand"),
    ("N_BAU9", "Purchase frequency"),
    ("N_BAU10", "Pack size purchased most often"),
]

PREFERRED_FILTER_FIELDS = [
    "starttime", "endtime", "today", "deviceid", "devicephonenum", "username", "simid",
    "caseid", "device_info", "duration", "City_1", "Sector", "Interviewer", "CS", "Q1a",
    "Age", "Age_Range", "Marital_Status", "Gender", "Sec_quest",
    "d3_q", "SEC", "Week",
    "PP_1", "PP_2", "PP_3", "PP", "randomdraw", "PP_1_1", "PP_1_4", "PP_1_2",
    "PP_1_5", "PP_1_3", "PP_1_6", "PP_1_7", "PP_1_8", "PP_1_9", "PP_1_11",
    "PP_1_10", "PP_1_16", "PP_1_14", "PP_1_15", "PP_1_12", "PP_1_13", "PP_2_2",
    "PP_2_5", "PP_2_1", "PP_2_3", "PP_2_4", "PP_2_6", "PP_2_13", "PP_2_11",
    "PP_2_7", "PP_2_9", "PP_2_16", "PP_2_10", "PP_2_8", "PP_2_15", "PP_2_14",
    "PP_2_12", "PP_3_5", "PP_3_3", "PP_3_2", "PP_3_4", "PP_3_1", "PP_3_16",
    "PP_3_13", "PP_3_15", "PP_3_11", "PP_3_10", "PP_3_8", "PP_3_7", "PP_3_9",
    "PP_3_12", "PP_3_14", "PP_3_6", "pp_note1", "pp_note2", "pp_note3",
    "Screener", "Intro1", "Consent1", "Consent2", "Intro2", "Consent3", "MAIN_INT",
    "NC", "Non_compete", "AG", "GE", "Age_cal", "S4_1", "SEC.grp", "SECn",
    "sumt", "E1", "E1_1", "E1_2", "E1_3", "E1_4", "E1_5", "E1_6", "E1_7",
    "E1_8", "E1_9", "E1_10", "E1_11", "E1_12", "E1_13", "E1_14", "E1_15",
    "E1_16", "E2", "E2_1", "E2_2", "E2_3", "E2_4", "E2_5", "E3", "E3_1",
    "E4", "E4_1", "E5", "E5_1", "E6", "E6_1", "E7", "E7_1", "E8", "E8_1",
]


def load_metadata(metadata_path: Path) -> dict[str, Any]:
    return json.loads(metadata_path.read_text(encoding="utf-8"))


def load_questionnaire_xlsform(workbook_path: Path) -> dict[str, Any]:
    """Read SurveyCTO's XLSForm directly into the decoder shape used by this service."""
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    if "survey" not in workbook.sheetnames or "choices" not in workbook.sheetnames:
        raise ValueError("Questionnaire workbook must contain survey and choices sheets")
    questions: dict[str, dict[str, str]] = {}
    survey_rows = workbook["survey"].iter_rows(values_only=True)
    survey_headers = [str(value or "").strip() for value in next(survey_rows)]
    survey_columns = {name: index for index, name in enumerate(survey_headers)}
    for row in survey_rows:
        name = str(row[survey_columns["name"]] or "").strip() if "name" in survey_columns else ""
        question_type = str(row[survey_columns["type"]] or "").strip() if "type" in survey_columns else ""
        if name and question_type:
            questions[name] = {
                "name": name,
                "type": question_type,
                "label": str(row[survey_columns["label"]] or "").strip() if "label" in survey_columns else name,
                "relevance": str(row[survey_columns["relevance"]] or "").strip() if "relevance" in survey_columns else "",
                "calculation": str(row[survey_columns["calculation"]] or "").strip() if "calculation" in survey_columns else "",
            }

    lists: dict[str, dict[str, str]] = {}
    choice_rows = workbook["choices"].iter_rows(values_only=True)
    choice_headers = [str(value or "").strip() for value in next(choice_rows)]
    choice_columns = {name: index for index, name in enumerate(choice_headers)}
    for row in choice_rows:
        list_name = str(row[choice_columns["list_name"]] or "").strip() if "list_name" in choice_columns else ""
        name = str(row[choice_columns["name"]] or "").strip() if "name" in choice_columns else ""
        if list_name and name:
            lists.setdefault(list_name, {})[name] = str(row[choice_columns["label"]] or name).strip()
    return {"questions": questions, "lists": lists}


def _clean_question_label(label: str) -> str:
    cleaned = re.sub(r"<[^>]+>", "", label)
    cleaned = re.sub(r"\s+", " ", cleaned)
    cleaned = cleaned.replace("\xa0", " ").strip()
    cleaned = re.sub(r"^([A-Za-z0-9_\.]+)\s*\.\s*", "", cleaned)
    cleaned = re.sub(r"^([A-Za-z0-9_\.]+)\s*:\s*", "", cleaned)
    cleaned = re.sub(r"^\s+", "", cleaned)
    cleaned = cleaned.strip()
    return cleaned


def _infer_parent_label(question_name: str, metadata: dict[str, Any], question_type: str | None = None) -> str:
    questions = metadata.get("questions", {})
    if not isinstance(questions, dict):
        return ""
    question = questions.get(question_name, {})
    if not isinstance(question, dict):
        return ""

    label = str(question.get("label") or "").strip()
    normalized_type = str(question_type or question.get("type", "") or "").strip().lower()
    if label and label.lower() not in {"others", "other"} and normalized_type != "calculate":
        cleaned = _clean_question_label(label)
        if cleaned and cleaned.lower() not in {question_name.lower(), "others", "other"}:
            return cleaned

    parent_candidates: list[str] = []
    for candidate in [question_name.rsplit("_OTH", 1)[0], re.sub(r"_OTH$", "", question_name)]:
        if candidate and candidate not in parent_candidates:
            parent_candidates.append(candidate)

    for candidate in parent_candidates:
        if candidate in questions and candidate != question_name:
            parent_label = str(questions[candidate].get("label") or "").strip()
            if parent_label:
                return _clean_question_label(parent_label)

    for candidate in [question_name, *parent_candidates]:
        if candidate in questions:
            parent_question = questions[candidate]
            if not isinstance(parent_question, dict):
                continue
            for referenced_name in re.findall(r"\$\{([A-Za-z0-9_.-]+)\}", str(parent_question.get("relevance", "") + " " + parent_question.get("calculation", "") + " " + parent_question.get("constraint", ""))):
                if referenced_name in questions:
                    referenced_label = str(questions[referenced_name].get("label") or "").strip()
                    if referenced_label:
                        cleaned_referenced = _clean_question_label(referenced_label)
                        if cleaned_referenced and cleaned_referenced.lower() not in {question_name.lower(), "others", "other"}:
                            return cleaned_referenced

    return ""


def _is_excluded_question_name(name: str, excluded_lookup: set[str]) -> bool:
    if str(name).lower() in excluded_lookup:
        return True

    normalized_name = str(name).strip()
    if not normalized_name:
        return False

    if re.fullmatch(r"pp(?:_\d+)*", normalized_name, flags=re.IGNORECASE):
        return True
    if re.fullmatch(r"pp_note\d+", normalized_name, flags=re.IGNORECASE):
        return True
    if re.fullmatch(r"e\d+(?:_\d+)?", normalized_name, flags=re.IGNORECASE):
        return True
    if re.fullmatch(r"sec(?:\.grp|n)?", normalized_name, flags=re.IGNORECASE):
        return True
    if re.fullmatch(r"intro\d+", normalized_name, flags=re.IGNORECASE):
        return True
    if re.fullmatch(r"consent\d+", normalized_name, flags=re.IGNORECASE):
        return True
    if normalized_name.lower() in {"screener", "main_int", "q1a", "randomdraw", "nc", "non_compete", "ag", "ge", "age", "age_cal", "age_range", "marital_status", "gender", "s4_1", "sumt"}:
        return True
    return False


def build_question_catalog(
    metadata: dict[str, Any],
    excluded_names: Sequence[str] | None = None,
    available_fields: Iterable[str] | None = None,
) -> list[tuple[str, str]]:
    """Create a question list for the sidebar from the XLSForm survey sheet.

    The sidebar should only contain analysis questions. Demographic and filter
    variables such as Gender, Sector, and starttime remain available for the
    filter/breakdown selector, but must not appear in the question dropdown.
    """
    questions = metadata.get("questions", {}) if isinstance(metadata, dict) else {}
    if not isinstance(questions, dict):
        return []

    excluded_lookup = {str(name).lower() for name in (excluded_names or PREFERRED_FILTER_FIELDS)}
    available_lookup = {str(name).lower() for name in available_fields} if available_fields is not None else None

    catalog: list[tuple[str, str]] = []
    seen: set[str] = set()
    for name, question in questions.items():
        # The current questionnaire can contain fields from later survey waves.
        # Do not offer a question when none of the imported submissions has it;
        # selecting one would otherwise always yield a zero-base table.
        if available_lookup is not None and str(name).lower() not in available_lookup:
            continue
        if _is_excluded_question_name(name, excluded_lookup):
            continue
        if not isinstance(question, dict):
            continue
        question_type = str(question.get("type", "")).strip().lower()
        # Only choice questions have response options that can be meaningfully
        # cross-tabulated by a filter.  Calculated flags (for example P1's 1/0)
        # and free-text fields must not appear in the question sidebar.
        if not question_type.startswith(("select_one", "select_multiple")):
            continue
        if question_type.startswith("note"):
            continue
        if question_type.startswith("begin") or question_type.startswith("end"):
            continue

        label = str(question.get("label") or "").strip()
        normalized = _clean_question_label(label)
        if not normalized:
            normalized = name

        if name.startswith("note_"):
            continue
        elif name.lower().startswith("randomdraw") or name.lower().startswith("panel"):
            normalized = normalized or name

        if not normalized and name in questions:
            normalized = str(questions[name].get("name") or name)

        if name not in seen:
            catalog.append((name, normalized))
            seen.add(name)
    return catalog


def load_template_registry(template_path: Path, metadata: dict[str, Any], category: str) -> list[tuple[str, str]]:
    """Extract the question sequence for a category from the report template's BackEnd sheet."""
    from openpyxl import load_workbook

    workbook = load_workbook(template_path, read_only=True, data_only=False)
    if "BackEnd" not in workbook.sheetnames:
        return table_registry(category)
    sheet = workbook["BackEnd"]
    category_column = None
    for row in sheet.iter_rows():
        for cell in row:
            if str(cell.value or "").strip().lower() == category.lower():
                category_column = cell.column
                break
        if category_column is not None:
            break
    if category_column is None:
        return table_registry(category)

    questions = metadata.get("questions", {})
    question_names_by_lower = {str(name).lower(): str(name) for name in questions}
    registry: list[tuple[str, str]] = []
    seen: set[str] = set()
    qbi_index = 0
    for row in sheet.iter_rows(min_col=category_column + 1, max_col=category_column + 1):
        text = str(row[0].value or "")
        match = re.match(r"\s*([A-Za-z][A-Za-z0-9_]*(?:\.\d+)?)", text)
        question_name = question_names_by_lower.get(match.group(1).lower(), "") if match else ""
        # The report's BackEnd tab has an older QFS spelling while the July
        # XLSForm and visible Noodles sheet use QFH.
        if not question_name and match and match.group(1).lower().startswith("n_qfs"):
            question_name = question_names_by_lower.get(("N_QFH" + match.group(1)[5:]).lower(), "")
        # The historic template uses a shared question stem for matrix / routed
        # items. Resolve those entries against the XLSForm rather than silently
        # dropping them from the report.
        if not question_name and match:
            stem = match.group(1)
            candidates = [
                name for name, question in questions.items()
                if str(name).lower().startswith(stem.lower() + ".") and str(question.get("type", "")).startswith("select_")
            ]
            if stem.lower() == "n_qbi":
                qbi_index += 1
                expected = f"N_QBI.{qbi_index}"
                question_name = question_names_by_lower.get(expected.lower(), "")
            elif stem.lower() == "n_bau4":
                lowered = text.lower()
                brand = re.search(r"about\s+(.+?)\s*$", text, flags=re.IGNORECASE)
                brand_name = brand.group(1).strip().lower() if brand else ""
                question_name = next((name for name in candidates if brand_name and brand_name in str(questions[name].get("label", "")).lower()), "")
        if question_name in questions and question_name not in seen:
            registry.append((question_name, text.strip()))
            seen.add(question_name)
    return registry or table_registry(category)


def _question(metadata: dict[str, Any], name: str) -> dict[str, Any]:
    questions = metadata.get("questions", {})
    if not isinstance(questions, dict):
        return {}
    exact = questions.get(name)
    if isinstance(exact, dict):
        return exact
    return next(
        (question for question_name, question in questions.items() if str(question_name).lower() == str(name).lower() and isinstance(question, dict)),
        {},
    )


def _options(metadata: dict[str, Any], name: str) -> dict[str, str]:
    question = _question(metadata, name)
    question_type = str(question.get("type", ""))
    list_name = question_type.replace("select_one", "", 1).replace("select_multiple", "", 1).strip()
    raw_options = metadata.get("lists", {}).get(list_name, {})
    return {str(key): str(label) for key, label in raw_options.items()}


def _choice_label(options: dict[str, str], value: Any) -> str:
    token = str(value).strip()
    if not token:
        return ""
    return options.get(token) or options.get(f"{token}.0") or token


def _answer_tokens(value: Any, question_type: str) -> list[Any]:
    if value in (None, ""):
        return []
    if str(question_type).startswith("select_multiple"):
        return value if isinstance(value, list) else str(value).split()
    return [value]


def _age_band(value: Any) -> str:
    """Match the three age bands used in the supplied full-data template."""
    try:
        age = int(float(value))
    except (TypeError, ValueError):
        return ""
    if 18 <= age <= 25:
        return "18 - 25 years"
    if 26 <= age <= 35:
        return "26 - 35 years"
    if 36 <= age <= 45:
        return "36 - 45 years"
    return ""


def _answers(payload: Any) -> dict[str, Any]:
    if not isinstance(payload, dict):
        return {}
    nested = payload.get("answers")
    return nested if isinstance(nested, dict) else payload


def _field_label(field_name: str, metadata: dict[str, Any]) -> str:
    question = _question(metadata, field_name)
    return str(question.get("label") or field_name) if question else field_name.replace("_", " ")


def available_filter_fields(payloads: Sequence[dict[str, Any]], metadata: dict[str, Any]) -> list[dict[str, str]]:
    """Return configured standard filters that are present in the submissions.

    Raw SurveyCTO payloads also contain thousands of dynamically generated keys;
    exposing every one would make the analyst selector unusable. Add new standard
    filter variables to PREFERRED_FILTER_FIELDS when a questionnaire requires it.
    """
    fields_by_lower: dict[str, str] = {}
    for payload in payloads:
        for key, value in _answers(payload).items():
            if value not in (None, ""):
                fields_by_lower[str(key).lower()] = str(key)
    preferred_by_lower = {field.lower(): field for field in PREFERRED_FILTER_FIELDS}
    ordered = [
        fields_by_lower[field.lower()]
        for field in PREFERRED_FILTER_FIELDS
        if field.lower() in fields_by_lower
    ]
    return [{"field": field, "label": _field_label(field, metadata)} for field in ordered]


def _resolve_field_name(field_name: str, fields: Sequence[str]) -> str:
    return next((field for field in fields if field.lower() == field_name.lower()), field_name)


def _row_value(row: dict[str, Any], field_name: str) -> Any:
    """Read a SurveyCTO field without assuming its key casing matches XLSForm."""
    if field_name in row:
        return row[field_name]
    resolved_name = _resolve_field_name(field_name, list(row))
    return row.get(resolved_name)


def _ordered_labels(options: dict[str, str], observed: Sequence[str]) -> list[str]:
    """Keep XLSForm choice order and include any non-standard submitted values."""
    labels = [label for label in options.values() if label]
    labels.extend(label for label in observed if label)
    return list(dict.fromkeys(labels))


def table_registry(category: str) -> list[tuple[str, str]]:
    if category.lower() != "noodles":
        return []
    return NOODLES_TABLES


def build_tables(
    payloads: Sequence[dict[str, Any]],
    metadata: dict[str, Any],
    category: str = "noodles",
    registry: Sequence[tuple[str, str]] | None = None,
    cut_fields: Sequence[str] | None = None,
    max_cut_groups: int = 50,
) -> list[dict[str, Any]]:
    """Calculate count and percentage tables for every configured question."""
    answer_rows = [_answers(payload) for payload in payloads]
    tables: list[dict[str, Any]] = []
    for question_name, title in registry or table_registry(category):
        question_name = next(
            (name for name in metadata.get("questions", {}) if str(name).lower() == question_name.lower()),
            question_name,
        )
        question = _question(metadata, question_name)
        if not question:
            continue
        question_type = str(question.get("type", ""))
        options = _options(metadata, question_name)
        valid_rows = [row for row in answer_rows if _row_value(row, question_name) not in (None, "")]
        total_base = len(valid_rows)
        counts: Counter[str] = Counter()
        for row in valid_rows:
            counts.update(_choice_label(options, token) for token in _answer_tokens(_row_value(row, question_name), question_type))

        cuts = []
        available_fields = list({key for row in answer_rows for key in row})
        requested_cut_fields = cut_fields or [
            field for field in PREFERRED_FILTER_FIELDS
            if any(available.lower() == field.lower() for available in available_fields)
        ]
        for requested_cut_name in requested_cut_fields:
            cut_name = _resolve_field_name(requested_cut_name, available_fields)
            if cut_name not in available_fields:
                continue
            cut_label = _field_label(cut_name, metadata)
            cut_options = _options(metadata, cut_name)
            groups: dict[str, list[dict[str, Any]]] = {}
            for row in valid_rows:
                if _row_value(row, cut_name) in (None, ""):
                    continue
                cut_type = str(_question(metadata, cut_name).get("type", ""))
                for token in _answer_tokens(_row_value(row, cut_name), cut_type):
                    label = _choice_label(cut_options, token)
                    if label:
                        groups.setdefault(label, []).append(row)
            # The filter's choices must be shown even when a choice has no
            # submissions yet.  For example, Gender always exposes the
            # XLSForm's Male and Female choices, not just the one seen first.
            group_labels = _ordered_labels(cut_options, list(groups))
            group_items = [(label, groups.get(label, [])) for label in group_labels]
            cuts.append(
                {
                    "field": cut_name,
                    "label": cut_label,
                    "groups": [
                        {
                            "label": group_label,
                            "base": len(group_rows),
                            "counts": dict(Counter(
                                _choice_label(options, token)
                                for row in group_rows
                                for token in _answer_tokens(_row_value(row, question_name), question_type)
                            )),
                        }
                        for group_label, group_rows in group_items[:max_cut_groups]
                    ],
                    "total_groups": len(group_items),
                    "truncated": len(group_items) > max_cut_groups,
                }
            )
        rows = [
            {"label": label, "count": counts[label], "pct": round((counts[label] / total_base) * 100, 1) if total_base else 0}
            for label in _ordered_labels(options, list(counts))
        ]
        tables.append({
            "id": question_name,
            "title": title,
            "question": str(question.get("label") or question_name),
            "base": total_base,
            "rows": rows,
            "cuts": cuts,
        })
    return tables
